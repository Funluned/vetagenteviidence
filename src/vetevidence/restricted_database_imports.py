"""Compliant, local-only imports for licensed and prediction data sources.

This module deliberately has no HTTP client.  GeneCards and MalaCards data may
only be imported after the caller confirms that the file was obtained under an
appropriate licence.  SwissTargetPrediction results are treated as
computational predictions, never as curated evidence.
"""

from __future__ import annotations

import csv
import math
import ntpath
import re
import unicodedata
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable
from urllib.parse import quote
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import load_workbook
from openpyxl.xml import DEFUSEDXML

from vetevidence.database_connectors import (
    AcquisitionMode,
    ConnectorResult,
    ConnectorStatus,
    DatabaseEvidenceClass,
    ProvenanceRecord,
    ResponseArtifact,
    canonical_json,
    sha256_bytes,
)


MAX_IMPORT_FILE_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 50_000
MAX_IMPORT_COLUMNS = 256
MAX_XLSX_ZIP_MEMBERS = 2_000
MAX_XLSX_EXPANDED_BYTES = 64 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200
_COMPRESSION_RATIO_MIN_BYTES = 4 * 1024 * 1024
OPENPYXL_DEFUSEDXML_ENABLED = bool(DEFUSEDXML)
_SUPPORTED_SUFFIXES = frozenset({".csv", ".tsv", ".xlsx"})
_MISSING_VALUES = frozenset(
    {"", "-", "n/a", "na", "none", "not available", "null", "未报告"}
)
_SAFE_CONTEXT_KEYS = frozenset(
    {
        "compound_id",
        "disease",
        "gene",
        "inchikey",
        "query",
        "smiles",
        "taxon_id",
    }
)
_SENSITIVE_CONTEXT_KEYS = frozenset(
    {
        "api_key",
        "apikey",
        "authorization",
        "email",
        "key",
        "license_confirmed",
        "password",
        "token",
    }
)
_MULTIVALUE_SEPARATOR = re.compile(r"[\n,;|]+")


@dataclass(frozen=True)
class _SourceSpec:
    key: str
    source_name: str
    endpoint_url: str
    license_url: str
    citation_url: str
    evidence_class: DatabaseEvidenceClass
    source_authenticity: str
    aliases: Mapping[str, Sequence[str]]
    required_columns: tuple[str, ...]
    parser: Callable[
        [Mapping[str, str], int, Mapping[str, Any]],
        dict[str, Any],
    ]


def _header_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).strip().casefold()
    normalized = normalized.replace("&", " and ")
    return " ".join(re.sub(r"[^0-9a-z]+", " ", normalized).split())


def _source_key(value: str) -> str:
    if not isinstance(value, str):
        raise ValueError("source_key must be a string.")
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _safe_filename(value: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError("filename cannot be blank.")
    if "\x00" in value:
        raise ValueError("filename contains a NUL character.")
    # ntpath recognizes both slash styles regardless of the host platform, so
    # Windows, POSIX, UNC, and mixed paths cannot leak directory components.
    filename = ntpath.basename(value.strip())
    if not filename or filename in {".", ".."}:
        raise ValueError("filename is invalid.")
    if len(filename) > 255:
        raise ValueError("filename is too long.")
    return filename


def _checked_payload(payload: bytes) -> bytes:
    if not isinstance(payload, bytes):
        raise ValueError("payload must be bytes.")
    if not payload:
        raise ValueError("The imported file is empty.")
    if len(payload) > MAX_IMPORT_FILE_BYTES:
        raise ValueError("The imported file exceeds the 10 MiB limit.")
    return payload


def _text_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("The imported table contains a non-finite number.")
        return format(value, ".15g")
    return str(value).strip()


def _inspect_xlsx_archive(payload: bytes) -> None:
    if not is_zipfile(BytesIO(payload)):
        raise ValueError("The .xlsx payload is not a valid OOXML archive.")
    try:
        with ZipFile(BytesIO(payload)) as archive:
            members = archive.infolist()
            names = {member.filename.replace("\\", "/") for member in members}
    except BadZipFile as exc:
        raise ValueError("The .xlsx payload is not a valid OOXML archive.") from exc

    if len(members) > MAX_XLSX_ZIP_MEMBERS:
        raise ValueError("The .xlsx archive contains too many members.")
    if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
        raise ValueError("The .xlsx archive is missing required workbook parts.")

    expanded_bytes = 0
    compressed_bytes = 0
    for member in members:
        normalized_name = member.filename.replace("\\", "/")
        path_parts = normalized_name.split("/")
        if (
            normalized_name.startswith("/")
            or ".." in path_parts
            or member.flag_bits & 0x1
        ):
            raise ValueError("The .xlsx archive contains an unsafe member.")
        lowered = normalized_name.casefold()
        if (
            lowered.endswith("vbaproject.bin")
            or lowered.startswith("xl/embeddings/")
            or lowered.startswith("xl/externallinks/")
        ):
            raise ValueError(
                "Macros, embedded objects, and external workbook links are not "
                "accepted."
            )
        expanded_bytes += member.file_size
        compressed_bytes += member.compress_size

    if expanded_bytes > MAX_XLSX_EXPANDED_BYTES:
        raise ValueError("The .xlsx archive expands beyond the 64 MiB limit.")
    if (
        expanded_bytes > _COMPRESSION_RATIO_MIN_BYTES
        and expanded_bytes
        > max(compressed_bytes, 1) * MAX_XLSX_COMPRESSION_RATIO
    ):
        raise ValueError("The .xlsx archive has a suspicious compression ratio.")


def _rows_from_delimited(
    payload: bytes,
    *,
    delimiter: str,
) -> list[list[str]]:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV and TSV imports must use UTF-8 encoding.") from exc
    if "\x00" in text:
        raise ValueError("The imported text contains a NUL character.")
    try:
        reader = csv.reader(StringIO(text, newline=""), delimiter=delimiter, strict=True)
        rows: list[list[str]] = []
        for row_number, raw_row in enumerate(reader, start=1):
            if len(raw_row) > MAX_IMPORT_COLUMNS:
                raise ValueError(
                    f"Row {row_number} exceeds the {MAX_IMPORT_COLUMNS}-column limit."
                )
            rows.append([value.strip() for value in raw_row])
            if len(rows) > MAX_IMPORT_ROWS + 1:
                raise ValueError(
                    f"The imported table exceeds {MAX_IMPORT_ROWS} data rows."
                )
    except csv.Error as exc:
        raise ValueError(f"The delimited file is malformed: {exc}.") from exc
    if not rows or not any(rows[0]):
        raise ValueError("The imported file does not contain a valid header row.")
    return rows


def _rows_from_xlsx(payload: bytes) -> list[list[str]]:
    if not OPENPYXL_DEFUSEDXML_ENABLED:
        raise RuntimeError(
            "XLSX imports require defusedxml-backed XML parsing; "
            "install the project dependencies before importing workbooks."
        )
    _inspect_xlsx_archive(payload)
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError("The .xlsx payload could not be parsed safely.") from exc

    try:
        populated_sheets: list[list[list[str]]] = []
        for sheet in workbook.worksheets:
            if (
                sheet.max_row > MAX_IMPORT_ROWS + 1
                or sheet.max_column > MAX_IMPORT_COLUMNS
            ):
                raise ValueError(
                    f"Worksheet {sheet.title!r} exceeds the table size limit."
                )
            sheet_rows: list[list[str]] = []
            for cells in sheet.iter_rows():
                values: list[str] = []
                for cell in cells:
                    if cell.data_type == "f":
                        raise ValueError(
                            "XLSX formulas are not accepted; replace them with "
                            "static values before import."
                        )
                    if cell.data_type == "e":
                        raise ValueError("XLSX error cells are not accepted.")
                    values.append(_text_value(cell.value))
                if any(values):
                    sheet_rows.append(values)
            if sheet_rows:
                populated_sheets.append(sheet_rows)
        if not populated_sheets:
            raise ValueError("The imported workbook is empty.")
        if len(populated_sheets) != 1:
            raise ValueError(
                "The imported workbook must contain exactly one non-empty worksheet."
            )
        return populated_sheets[0]
    finally:
        workbook.close()


def _read_rows(filename: str, payload: bytes) -> tuple[list[list[str]], str]:
    suffix = Path(filename).suffix.casefold()
    if suffix not in _SUPPORTED_SUFFIXES:
        raise ValueError("Only CSV, TSV, and XLSX imports are supported.")
    if suffix == ".csv":
        return _rows_from_delimited(payload, delimiter=","), (
            "text/csv; charset=utf-8"
        )
    if suffix == ".tsv":
        return _rows_from_delimited(payload, delimiter="\t"), (
            "text/tab-separated-values; charset=utf-8"
        )
    return _rows_from_xlsx(payload), (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def _alias_index(spec: _SourceSpec) -> dict[str, str]:
    index: dict[str, str] = {}
    for canonical, aliases in spec.aliases.items():
        for alias in (canonical, *aliases):
            normalized = _header_key(alias)
            existing = index.get(normalized)
            if existing is not None and existing != canonical:
                raise RuntimeError(
                    f"Header alias {alias!r} is ambiguous in the source schema."
                )
            index[normalized] = canonical
    return index


def _canonicalize_rows(
    raw_rows: list[list[str]],
    *,
    spec: _SourceSpec,
) -> list[tuple[int, dict[str, str]]]:
    raw_headers = raw_rows[0]
    if not raw_headers or any(not header.strip() for header in raw_headers):
        raise ValueError("The header row contains an empty column name.")
    for header in raw_headers:
        if len(header) > 256 or any(
            unicodedata.category(character) == "Cc" for character in header
        ):
            raise ValueError("The header row contains an illegal column name.")

    normalized_headers = [_header_key(header) for header in raw_headers]
    if any(not value for value in normalized_headers):
        raise ValueError("The header row contains an illegal column name.")
    duplicate_raw = sorted(
        {
            raw_headers[index]
            for index, value in enumerate(normalized_headers)
            if normalized_headers.count(value) > 1
        }
    )
    if duplicate_raw:
        raise ValueError(
            "The header row contains duplicate columns: "
            + ", ".join(duplicate_raw)
            + "."
        )

    aliases = _alias_index(spec)
    canonical_headers = [aliases.get(value) for value in normalized_headers]
    duplicate_canonical = sorted(
        {
            canonical
            for canonical in canonical_headers
            if canonical is not None and canonical_headers.count(canonical) > 1
        }
    )
    if duplicate_canonical:
        raise ValueError(
            "Multiple columns map ambiguously to: "
            + ", ".join(duplicate_canonical)
            + "."
        )
    missing = [
        column
        for column in spec.required_columns
        if column not in canonical_headers
    ]
    if missing:
        raise ValueError(
            f"{spec.source_name} import is missing required columns: "
            + ", ".join(missing)
            + "."
        )

    canonical_rows: list[tuple[int, dict[str, str]]] = []
    for row_number, values in enumerate(raw_rows[1:], start=2):
        if len(values) != len(raw_headers):
            raise ValueError(
                f"Row {row_number} has {len(values)} columns; "
                f"{len(raw_headers)} were expected."
            )
        if not any(values):
            continue
        row = {
            canonical: values[index].strip()
            for index, canonical in enumerate(canonical_headers)
            if canonical is not None
        }
        missing_values = [
            column for column in spec.required_columns if _is_missing(row[column])
        ]
        if missing_values:
            raise ValueError(
                f"Row {row_number} has blank required values: "
                + ", ".join(missing_values)
                + "."
            )
        canonical_rows.append((row_number, row))
        if len(canonical_rows) > MAX_IMPORT_ROWS:
            raise ValueError(
                f"The imported table exceeds {MAX_IMPORT_ROWS} data rows."
            )
    return canonical_rows


def _is_missing(value: str | None) -> bool:
    return value is None or value.strip().casefold() in _MISSING_VALUES


def _optional_text(value: str | None) -> str | None:
    if _is_missing(value):
        return None
    assert value is not None
    return value.strip()


def _required_text(value: str | None, *, label: str, row_number: int) -> str:
    cleaned = _optional_text(value)
    if cleaned is None:
        raise ValueError(f"Row {row_number} is missing {label}.")
    return cleaned


def _optional_float(
    value: str | None,
    *,
    label: str,
    row_number: int,
) -> float | None:
    cleaned = _optional_text(value)
    if cleaned is None:
        return None
    try:
        parsed = float(cleaned.replace(",", ""))
    except ValueError as exc:
        raise ValueError(
            f"Row {row_number} has an invalid {label}: {cleaned!r}."
        ) from exc
    if not math.isfinite(parsed):
        raise ValueError(f"Row {row_number} has a non-finite {label}.")
    return parsed


def _probability(value: str | None, *, row_number: int) -> float:
    cleaned = _required_text(value, label="probability", row_number=row_number)
    percentage = cleaned.endswith("%")
    number_text = cleaned[:-1].strip() if percentage else cleaned
    try:
        parsed = float(number_text)
    except ValueError as exc:
        raise ValueError(
            f"Row {row_number} has an invalid probability: {cleaned!r}."
        ) from exc
    if not math.isfinite(parsed):
        raise ValueError(f"Row {row_number} has a non-finite probability.")
    if percentage:
        parsed /= 100
    if parsed < 0 or parsed > 1:
        raise ValueError(
            f"Row {row_number} probability must be between 0 and 1."
        )
    return parsed


def _known_actives(value: str | None, *, row_number: int) -> int | str | None:
    cleaned = _optional_text(value)
    if cleaned is None:
        return None
    if re.fullmatch(r"\d+", cleaned):
        return int(cleaned)
    if re.fullmatch(r"\d+\s*/\s*\d+", cleaned):
        return re.sub(r"\s+", "", cleaned)
    raise ValueError(
        f"Row {row_number} has an invalid known-actives value: {cleaned!r}."
    )


def _split_values(value: str | None) -> list[str]:
    cleaned = _optional_text(value)
    if cleaned is None:
        return []
    return list(
        dict.fromkeys(
            part.strip()
            for part in _MULTIVALUE_SEPARATOR.split(cleaned)
            if part.strip()
        )
    )


def _split_identifier_values(value: str | None) -> list[str]:
    cleaned = _optional_text(value)
    if cleaned is None:
        return []
    return list(
        dict.fromkeys(
            part.strip()
            for part in re.split(r"[\s,;|]+", cleaned)
            if part.strip()
        )
    )


def _xrefs(
    row: Mapping[str, str],
    namespaces: Mapping[str, str],
) -> dict[str, list[str]]:
    values: dict[str, list[str]] = {}
    for column, namespace in namespaces.items():
        identifiers = (
            _split_values(row.get(column))
            if namespace == "aliases"
            else _split_identifier_values(row.get(column))
        )
        if identifiers:
            values[namespace] = list(
                dict.fromkeys((*values.get(namespace, []), *identifiers))
            )
    generic = _split_values(row.get("xrefs"))
    if generic:
        values["other"] = generic
    return values


_GENECARDS_XREFS = {
    "xref_genecards": "genecards",
    "xref_hgnc": "hgnc",
    "xref_entrez": "entrez_gene",
    "xref_ensembl": "ensembl_gene",
    "xref_uniprot": "uniprot",
    "xref_uniprot_swissprot": "uniprot",
    "xref_uniprot_trembl": "uniprot",
    "xref_omim": "omim",
    "xref_aliases": "aliases",
}
_MALACARDS_XREFS = {
    "xref_omim": "omim",
    "xref_orpha": "orphanet",
    "xref_umls": "umls",
    "xref_icd10": "icd10",
    "xref_mesh": "mesh",
}


def _parse_genecards(
    row: Mapping[str, str],
    row_number: int,
    _: Mapping[str, Any],
) -> dict[str, Any]:
    symbol = _required_text(row.get("symbol"), label="symbol", row_number=row_number)
    return {
        "record_type": "genecards_gene",
        "symbol": symbol,
        "name": _optional_text(row.get("name")),
        "gene_type": _optional_text(row.get("gene_type")),
        "relevance_score": _optional_float(
            row.get("relevance_score"),
            label="relevance score",
            row_number=row_number,
        ),
        "knowledge_score": _optional_float(
            row.get("knowledge_score"),
            label="knowledge score",
            row_number=row_number,
        ),
        "xrefs": _xrefs(row, _GENECARDS_XREFS),
        "taxon_id": 9606,
        "source_url": (
            "https://www.genecards.org/cgi-bin/carddisp.pl?gene="
            + quote(symbol, safe="")
        ),
    }


def _parse_malacards(
    row: Mapping[str, str],
    row_number: int,
    _: Mapping[str, Any],
) -> dict[str, Any]:
    mcid = _required_text(row.get("mcid"), label="MCID", row_number=row_number)
    disease_name = _required_text(
        row.get("disease_name"),
        label="disease name",
        row_number=row_number,
    )
    return {
        "record_type": "malacards_disease",
        "mcid": mcid,
        "disease_name": disease_name,
        "family": _optional_text(row.get("family")),
        "mifts_score": _optional_float(
            row.get("mifts_score"),
            label="MIFTS score",
            row_number=row_number,
        ),
        "relevance_score": _optional_float(
            row.get("relevance_score"),
            label="relevance score",
            row_number=row_number,
        ),
        "xrefs": _xrefs(row, _MALACARDS_XREFS),
        "taxon_id": 9606,
        "source_url": (
            "https://www.malacards.org/card/"
            + quote(mcid.casefold(), safe="_-")
        ),
    }


def _parse_swiss_target_prediction(
    row: Mapping[str, str],
    row_number: int,
    query_context: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "record_type": "swiss_target_prediction",
        "target_name": _required_text(
            row.get("target_name"),
            label="target name",
            row_number=row_number,
        ),
        "gene_names": _split_identifier_values(row.get("gene_names")),
        "uniprot_ids": _split_identifier_values(row.get("uniprot_ids")),
        "chembl_ids": _split_identifier_values(row.get("chembl_ids")),
        "target_class": _optional_text(row.get("target_class")),
        "probability": _probability(
            row.get("probability"),
            row_number=row_number,
        ),
        "known_actives": _known_actives(
            row.get("known_actives"),
            row_number=row_number,
        ),
        "query_smiles": query_context["smiles"],
        "taxon_id": query_context["taxon_id"],
        "source_url": "https://www.swisstargetprediction.ch/",
    }


_GENECARDS_ALIASES: Mapping[str, Sequence[str]] = {
    "symbol": (
        "Gene Symbol",
        "GeneCards Symbol",
        "HGNC Symbol",
        "Approved Symbol",
    ),
    "name": (
        "Gene Name",
        "Gene Description",
        "Description",
        "Gene Name/Description",
    ),
    "gene_type": ("Gene Type", "Gene Category", "Category", "Type"),
    "relevance_score": (
        "Relevance",
        "Relevance Score",
        "GeneCards Relevance Score",
    ),
    "knowledge_score": (
        "Knowledge Score",
        "GeneCards Knowledge Score",
        "GIFtS",
        "GeneCards Inferred Functionality Score",
        "GeneCards Inferred Functionality Score (GIFtS)",
    ),
    "xrefs": ("Cross References", "Cross-References", "External IDs"),
    "xref_genecards": ("GeneCards ID", "GeneCard ID", "GC ID"),
    "xref_hgnc": ("HGNC", "HGNC ID", "HGNC IDs"),
    "xref_entrez": (
        "Entrez",
        "Entrez ID",
        "Entrez Gene",
        "Entrez Gene ID",
        "NCBI Gene ID",
    ),
    "xref_ensembl": (
        "Ensembl",
        "Ensembl ID",
        "Ensembl Gene",
        "Ensembl Gene ID",
    ),
    "xref_uniprot": (
        "UniProt",
        "UniProt ID",
        "UniProt IDs",
        "UniProtKB",
    ),
    "xref_uniprot_swissprot": (
        "UniProtKB/Swiss-Prot ID",
        "UniProtKB Swiss-Prot ID",
        "Swiss-Prot ID",
    ),
    "xref_uniprot_trembl": (
        "UniProtKB/TrEMBL ID",
        "UniProtKB TrEMBL ID",
        "TrEMBL ID",
    ),
    "xref_omim": ("OMIM", "OMIM ID", "MIM", "MIM Number"),
    "xref_aliases": (
        "Aliases",
        "Gene Aliases",
        "Synonyms",
        "Alias Symbols",
    ),
}
_MALACARDS_ALIASES: Mapping[str, Sequence[str]] = {
    "mcid": (
        "MCID",
        "MalaCards ID",
        "MalaCard ID",
        "MalaCards Identifier",
    ),
    "disease_name": ("Disease Name", "Disease", "Name"),
    "family": ("Disease Family", "Family"),
    "mifts_score": ("MIFTS", "MIFTS Score"),
    "relevance_score": ("Relevance", "Relevance Score"),
    "xrefs": ("Cross References", "Cross-References", "External IDs"),
    "xref_omim": ("OMIM", "OMIM ID", "MIM", "MIM Number"),
    "xref_orpha": (
        "ORPHA",
        "ORPHA ID",
        "Orphanet",
        "Orphanet ID",
    ),
    "xref_umls": ("UMLS", "UMLS ID", "UMLS CUI"),
    "xref_icd10": ("ICD10", "ICD-10", "ICD 10"),
    "xref_mesh": ("MeSH", "MeSH ID"),
}
_SWISS_TARGET_ALIASES: Mapping[str, Sequence[str]] = {
    "target_name": (
        "Target",
        "Target Name",
        "Predicted Target",
        "Protein Target",
    ),
    "gene_names": (
        "Gene",
        "Gene Name",
        "Gene Names",
        "Common Name",
        "Common Names",
    ),
    "uniprot_ids": (
        "UniProt",
        "UniProt ID",
        "UniProt IDs",
        "UniProtKB",
    ),
    "chembl_ids": ("ChEMBL", "ChEMBL ID", "ChEMBL IDs"),
    "target_class": ("Target Class", "Class"),
    "probability": (
        "Probability",
        "Probability*",
        "Prediction Probability",
        "Score",
    ),
    "known_actives": (
        "Known Actives",
        "Known Actives (3D/2D)",
        "Known Actives 3D/2D",
    ),
}


_SOURCE_SPECS = {
    "genecards": _SourceSpec(
        key="genecards",
        source_name="GeneCards",
        endpoint_url="https://www.genecards.org/",
        license_url="https://www.lifemapsc.com/terms-of-use/",
        citation_url="https://www.genecards.org/Guide/GeneCard",
        evidence_class=DatabaseEvidenceClass.CURATED_DATABASE,
        source_authenticity="user_attested_licensed_export",
        aliases=_GENECARDS_ALIASES,
        required_columns=("symbol",),
        parser=_parse_genecards,
    ),
    "malacards": _SourceSpec(
        key="malacards",
        source_name="MalaCards",
        endpoint_url="https://www.malacards.org/",
        license_url="https://www.lifemapsc.com/terms-of-use/",
        citation_url="https://www.malacards.org/pages/publications",
        evidence_class=DatabaseEvidenceClass.CURATED_DATABASE,
        source_authenticity="user_attested_licensed_export",
        aliases=_MALACARDS_ALIASES,
        required_columns=("mcid", "disease_name"),
        parser=_parse_malacards,
    ),
    "swisstargetprediction": _SourceSpec(
        key="swisstargetprediction",
        source_name="SwissTargetPrediction",
        endpoint_url="https://www.swisstargetprediction.ch/",
        license_url="https://www.swisstargetprediction.ch/termsofuse.php",
        citation_url="https://www.swisstargetprediction.ch/about.php",
        evidence_class=DatabaseEvidenceClass.COMPUTATIONAL_PREDICTION,
        source_authenticity="user_attested_manual_prediction",
        aliases=_SWISS_TARGET_ALIASES,
        required_columns=("target_name", "probability"),
        parser=_parse_swiss_target_prediction,
    ),
}
_SOURCE_KEY_ALIASES = {
    "genecard": "genecards",
    "genecards": "genecards",
    "malacard": "malacards",
    "malacards": "malacards",
    "swisstarget": "swisstargetprediction",
    "swisstargetprediction": "swisstargetprediction",
}


def _validated_query_context(
    spec: _SourceSpec,
    query_context: Mapping[str, Any] | None,
) -> dict[str, Any]:
    if query_context is None:
        incoming: Mapping[str, Any] = {}
    elif isinstance(query_context, Mapping):
        incoming = query_context
    else:
        raise ValueError("query_context must be a mapping.")

    if spec.key == "swisstargetprediction":
        smiles_value = incoming.get("smiles")
        if not isinstance(smiles_value, str) or not smiles_value.strip():
            raise ValueError(
                "SwissTargetPrediction import requires query_context.smiles."
            )
        smiles = smiles_value.strip()
        if len(smiles) > 10_000 or any(
            unicodedata.category(character) == "Cc" for character in smiles
        ):
            raise ValueError("query_context.smiles is invalid.")
        taxon_value = incoming.get("taxon_id")
        if isinstance(taxon_value, bool):
            raise ValueError("query_context.taxon_id is invalid.")
        try:
            taxon_id = int(taxon_value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                "SwissTargetPrediction import requires query_context.taxon_id."
            ) from exc
        if taxon_id not in {9606, 10090, 10116}:
            raise ValueError(
                "SwissTargetPrediction taxon_id must be 9606, 10090, or 10116."
            )
        safe: dict[str, Any] = {"smiles": smiles, "taxon_id": taxon_id}
        inchikey = incoming.get("inchikey")
        if isinstance(inchikey, str) and inchikey.strip():
            safe["inchikey"] = inchikey.strip()
        return safe

    safe = {}
    for key, value in incoming.items():
        normalized_key = str(key).strip().casefold()
        if (
            normalized_key in _SENSITIVE_CONTEXT_KEYS
            or normalized_key not in _SAFE_CONTEXT_KEYS
        ):
            continue
        if normalized_key == "taxon_id":
            if isinstance(value, bool):
                raise ValueError(f"{spec.source_name} taxon_id must be 9606.")
            try:
                taxon_id = int(value)
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"{spec.source_name} taxon_id must be 9606."
                ) from exc
            if taxon_id != 9606:
                raise ValueError(f"{spec.source_name} taxon_id must be 9606.")
            safe[normalized_key] = taxon_id
            continue
        if isinstance(value, float) and not math.isfinite(value):
            continue
        if isinstance(value, (str, int, float, bool)) and len(str(value)) <= 2_000:
            rendered = str(value)
            if any(
                unicodedata.category(character) == "Cc"
                for character in rendered
            ):
                continue
            safe[normalized_key] = value
    return safe


def _stable_ids_for(
    spec: _SourceSpec,
    records: Sequence[Mapping[str, Any]],
    query_context: Mapping[str, Any],
) -> tuple[str, ...]:
    stable_ids: list[str] = []
    if spec.key in {"genecards", "malacards"}:
        stable_ids.append("NCBITaxon:9606")
    elif spec.key == "swisstargetprediction":
        stable_ids.append(f"NCBITaxon:{query_context['taxon_id']}")
        inchikey = query_context.get("inchikey")
        if inchikey:
            stable_ids.append(f"InChIKey:{inchikey}")

    for record in records:
        if spec.key == "genecards":
            stable_ids.append(f"GeneCards:{record['symbol']}")
            stable_ids.extend(
                _qualified_xrefs(
                    record["xrefs"],
                    {
                        "genecards": "GeneCards",
                        "hgnc": "HGNC",
                        "entrez_gene": "NCBIGene",
                        "ensembl_gene": "Ensembl",
                        "uniprot": "UniProtKB",
                        "omim": "OMIM",
                    },
                )
            )
        elif spec.key == "malacards":
            stable_ids.append(f"MalaCards:{record['mcid']}")
            stable_ids.extend(
                _qualified_xrefs(
                    record["xrefs"],
                    {
                        "omim": "OMIM",
                        "orphanet": "Orphanet",
                        "umls": "UMLS",
                        "icd10": "ICD10",
                        "mesh": "MeSH",
                    },
                )
            )
        else:
            stable_ids.extend(
                f"UniProtKB:{identifier}"
                for identifier in record["uniprot_ids"]
            )
            stable_ids.extend(
                f"ChEMBL:{identifier}" for identifier in record["chembl_ids"]
            )
    return tuple(dict.fromkeys(stable_ids))


def _qualified_xrefs(
    xrefs: Mapping[str, Sequence[str]],
    prefixes: Mapping[str, str],
) -> list[str]:
    stable_ids: list[str] = []
    for namespace, prefix in prefixes.items():
        for value in xrefs.get(namespace, ()):
            if value.casefold().startswith(prefix.casefold() + ":"):
                stable_ids.append(value)
            else:
                stable_ids.append(f"{prefix}:{value}")
    return stable_ids


def import_restricted_database_file(
    source_key: str,
    filename: str,
    payload: bytes,
    *,
    license_confirmed: bool,
    query_context: Mapping[str, Any] | None = None,
    source_version: str | None = None,
) -> ConnectorResult:
    """Import one licensed database export or prediction result without I/O.

    The function reads only ``payload`` and never opens ``filename`` or makes a
    network request.  A syntactically valid header-only file is archived and
    returns ``NO_RESULTS``; a truly empty or malformed file is rejected.
    """

    normalized_source = _SOURCE_KEY_ALIASES.get(_source_key(source_key))
    if normalized_source is None:
        raise ValueError(
            "source_key must be GeneCards, MalaCards, or SwissTargetPrediction."
        )
    spec = _SOURCE_SPECS[normalized_source]
    if license_confirmed is not True:
        if spec.key in {"genecards", "malacards"}:
            raise ValueError(
                f"{spec.source_name} import requires explicit license "
                "confirmation."
            )
        raise ValueError(
            "SwissTargetPrediction import requires explicit confirmation "
            "that the result was generated manually for permitted use."
        )

    safe_name = _safe_filename(filename)
    raw = _checked_payload(payload)
    raw_rows, content_type = _read_rows(safe_name, raw)
    safe_context = _validated_query_context(spec, query_context)
    canonical_rows = _canonicalize_rows(raw_rows, spec=spec)
    records = tuple(
        {
            **spec.parser(row, row_number, safe_context),
            "source_authenticity": spec.source_authenticity,
        }
        for row_number, row in canonical_rows
    )

    payload_sha256 = sha256_bytes(raw)
    normalized = canonical_json(
        {
            "byte_count": len(raw),
            "filename": safe_name,
            "query_context": safe_context,
            "sha256": payload_sha256,
            "source_attestation": {
                "confirmed": True,
                "terms_url": spec.license_url,
                "type": spec.source_authenticity,
            },
        }
    )
    cleaned_source_version = _optional_text(source_version)
    if cleaned_source_version is not None and (
        len(cleaned_source_version) > 256
        or any(
            unicodedata.category(character) == "Cc"
            for character in cleaned_source_version
        )
    ):
        raise ValueError("source_version is invalid.")
    provenance = ProvenanceRecord(
        source_name=spec.source_name,
        endpoint_url=spec.endpoint_url,
        method="IMPORT",
        normalized_request=normalized,
        request_sha256=sha256_bytes(normalized.encode("utf-8")),
        raw_response_sha256=payload_sha256,
        retrieved_at_utc=datetime.now(UTC),
        http_status=None,
        content_type=content_type,
        source_version=cleaned_source_version,
        stable_ids=_stable_ids_for(spec, records, safe_context),
        license_url=spec.license_url,
        citation_url=spec.citation_url,
    )
    artifact = ResponseArtifact(
        provenance=provenance,
        raw_response=raw,
    )
    status = ConnectorStatus.OK if records else ConnectorStatus.NO_RESULTS
    warnings = (
        ()
        if records
        else ("The imported file contains headers but no data records.",)
    )
    return ConnectorResult(
        status=status,
        acquisition_mode=AcquisitionMode.MANUAL_IMPORT,
        evidence_class=spec.evidence_class,
        records=records,
        artifacts=(artifact,),
        warnings=warnings,
    )


__all__ = [
    "MAX_IMPORT_FILE_BYTES",
    "import_restricted_database_file",
]
